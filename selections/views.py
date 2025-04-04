import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from django.http import HttpResponse
from .models import Selection

def download_csv_tier_report(request, tier_code):
    """
    Gera e faz o download do relatório CSV de uma turma específica.
    """
    tiers = dict(Selection.TIER_CHOICES)

    # Verifica se o código da turma é válido
    if tier_code not in tiers:
        return HttpResponse("Turma inválida.", status=400)

    selections = Selection.objects.filter(tier=tier_code).select_related('teacher')

    # Configura a resposta HTTP para download do CSV
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{tier_code}.csv"'

    writer = csv.writer(response)
    
    # Escreve o cabeçalho do CSV
    writer.writerow(["Nome do Aluno", "Email", "Número de Telefone" ,"Turma", "Professor"])

    # Escreve os dados das seleções
    for selection in selections:
        writer.writerow([
            selection.student_name,
            selection.email,
            selection.phone_number,
            tiers[tier_code],  # Nome completo da turma
            selection.teacher.name
        ])

    return response


def download_excel_tier_report(request):
    tiers = dict(Selection.TIER_CHOICES)

    selections = Selection.objects.select_related('teacher').order_by('tier')

    # Criar um arquivo Excel
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = f"Relatório"  # Nome da aba no Excel

    # Estilização do cabeçalho
    headers = ["Nome do Aluno", "Email", "Número de Telefone", "Turma", "Professor"]
    sheet.append(headers)

    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for cell in sheet[1]:  
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Estilização das linhas (cor alternada e bordas)
    row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    previous_tier = None
    row_index = 2  # Começa a partir da linha 2 para as seleções
    for selection in selections:
        # Verifica se a turma mudou, se sim, adiciona uma linha em branco
        if selection.tier != previous_tier:
            if previous_tier is not None:  # Para evitar linha extra no início
                sheet.append([])  # Linha em branco entre turmas
                row_index += 1  # Avança a linha para a próxima turma
            previous_tier = selection.tier

        # Adiciona os dados da seleção
        row_data = [
            selection.student_name,
            selection.email,
            selection.phone_number,
            selection.tier,
            selection.teacher.name
        ]
        sheet.append(row_data)

        # Estilização da linha (bordas e cor alternada)
        for cell in sheet[row_index]:
            cell.border = thin_border
            if row_index % 2 == 0:  # Aplica fundo alternado nas linhas pares
                cell.fill = row_fill

        row_index += 1  # Avança para a próxima linha

    # Ajustar automaticamente a largura das colunas
    for col in sheet.columns:
        max_length = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        sheet.column_dimensions[col_letter].width = max_length + 5  # Adiciona um espaço extra

    # Configurar a resposta HTTP para download do Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Relatório.xlsx"'

    # Salvar o arquivo Excel na resposta HTTP
    workbook.save(response)

    return response
